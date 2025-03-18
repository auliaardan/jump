import json
import datetime
import re

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.mail import EmailMessage
from django.core.paginator import Paginator
from django.db.models import Sum
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.template.loader import render_to_string
from django.utils import timezone
from django.views import View
from django.views.decorators.http import require_POST
from django.views.generic import ListView, DetailView
from openpyxl.workbook import Workbook

from .forms import PaymentProofForm, UserRegisterForm, SciComSubmissionForm
from .models import (
    PaymentMethod, Seminar, Order, landing_page, Cart, CartItem, about_us, seminars_page,
    workshops_page, DiscountCode, PaymentProof, scicom_rules, qrcode, ImageForPage, Sponsor,
    SciComSubmission, TicketCategory, OrderItem
)


# Helper functions and mixins

def aggregate_order_items(order):
    """Aggregate order items by seminar and ticket category."""
    items = {}
    for item in order.orderitem_set.all():
        seminar = item.ticket_category.seminar
        ticket_category = item.ticket_category
        key = (seminar.id, ticket_category.id)
        if key in items:
            items[key]['quantity'] += item.quantity
        else:
            items[key] = {
                'seminar': seminar,
                'ticket_category': ticket_category,
                'quantity': item.quantity,
            }
    return list(items.values())


def aggregate_order_items_str(order):
    """Return a string describing aggregated order items."""
    items = {}
    for item in order.orderitem_set.all():
        seminar = item.ticket_category.seminar
        ticket_category = item.ticket_category
        key = (seminar.id, ticket_category.id)
        if key in items:
            items[key]['quantity'] += item.quantity
        else:
            items[key] = {
                'seminar_title': seminar.title,
                'seminar_date': seminar.date.strftime('%Y-%m-%d'),
                'ticket_category_name': ticket_category.name,
                'quantity': item.quantity,
            }
    return "; ".join(
        f"{v['seminar_title']} ({v['seminar_date']}) - {v['ticket_category_name']} - Quantity: {v['quantity']}"
        for v in items.values()
    )


def send_html_email(subject, template, context, recipient):
    """Helper function to send an HTML email."""
    body = render_to_string(template, context)
    email = EmailMessage(subject, body, 'admin@jump2025.com', [recipient])
    email.content_subtype = 'html'
    email.send()


class SeminarPaginationMixin:
    paginate_by = 4

    def paginate_seminars(self, seminars):
        paginator = Paginator(seminars, self.paginate_by)
        page_number = self.request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        num_placeholders = self.paginate_by - len(page_obj) if len(page_obj) < self.paginate_by else 0
        return {
            'seminar_list': page_obj,
            'num_placeholders': num_placeholders,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'search_query': self.request.GET.get('search', '')
        }


# Views

@login_required
def create_submission(request):
    if request.method == 'POST':
        form = SciComSubmissionForm(request.POST)
        if form.is_valid():
            submission = form.save(commit=False)
            submission.user = request.user
            submission.save()
            send_html_email(
                'Your Submission Confirmation',
                'tickets/emails/submission_confirmed.html',
                {
                    'name': submission.user.nama_lengkap,
                    'type': submission.get_submission_type_display(),
                },
                submission.user.email
            )
            return redirect('seminar_list')
    else:
        form = SciComSubmissionForm(initial={'submission_type': 'abstract'})
        form.instance.user = request.user

    return render(request, 'tickets/create_submission.html', {'form': form})


class SponsorsView(ListView):
    model = Sponsor
    template_name = 'Sponsors.html'
    context_object_name = 'sponsors'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sponsors = Sponsor.objects.all()
        context.update({
            'platinum_sponsors': sponsors.filter(category='Large'),
            'gold_sponsors': sponsors.filter(category='Medium'),
            'silver_sponsors': sponsors.filter(category='Small'),
        })
        return context


class ScicomView(SeminarPaginationMixin, ListView):
    model = scicom_rules
    template_name = 'scicom.html'
    context_object_name = 'scicom_rules'

    def get_queryset(self):
        return scicom_rules.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        search = self.request.GET.get('search', '')
        seminars = Seminar.objects.filter(title__icontains=search).order_by('date') if search \
            else Seminar.objects.all().order_by('date')
        context.update(self.paginate_seminars(seminars))
        context['qrcode'] = qrcode.objects.last()
        context['images'] = ImageForPage.objects.filter(category=ImageForPage.SCICOM)
        return context


class WorkshopView(SeminarPaginationMixin, ListView):
    model = workshops_page
    template_name = 'Workshops.html'
    context_object_name = 'workshops'

    def get_queryset(self):
        return workshops_page.objects.last()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        search = self.request.GET.get('search', '')
        seminars = Seminar.objects.filter(title__icontains=search).order_by('date') if search \
            else Seminar.objects.filter(category=Seminar.WORKSHOP).order_by('date')
        context.update(self.paginate_seminars(seminars))
        context['images'] = ImageForPage.objects.filter(category=ImageForPage.WORKSHOP)
        return context


class SeminarsView(SeminarPaginationMixin, ListView):
    model = seminars_page
    template_name = 'Seminars.html'
    context_object_name = 'seminars'

    def get_queryset(self):
        return seminars_page.objects.last()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        search = self.request.GET.get('search', '')
        seminars = Seminar.objects.filter(title__icontains=search).order_by('date') if search \
            else Seminar.objects.filter(category=Seminar.SEMINAR).order_by('date')
        context.update(self.paginate_seminars(seminars))
        context['images'] = ImageForPage.objects.filter(category=ImageForPage.SEMINAR)
        return context


def coming_soon_view(request):
    return render(request, 'coming_soon.html')


class BaseView(SeminarPaginationMixin, ListView):
    model = landing_page
    template_name = 'index.html'
    context_object_name = "landing"

    def get_queryset(self):
        return landing_page.objects.last()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        search = self.request.GET.get('search', '')
        seminars = Seminar.objects.filter(title__icontains=search).order_by('date') if search \
            else Seminar.objects.all().order_by('date')
        context.update(self.paginate_seminars(seminars))
        context['seminars_all'] = Seminar.objects.all().order_by('date')
        now = timezone.now()
        next_seminar = Seminar.objects.filter(date__gte=now).order_by('date').first()
        context['next_seminar'] = next_seminar
        context['all_events_over'] = not next_seminar
        return context


class AboutUsView(ListView):
    model = about_us
    template_name = 'about_us.html'
    context_object_name = "about_us"

    def get_queryset(self):
        return about_us.objects.last()


def order_confirmed(request):
    return render(request, 'tickets/order_confirmed.html')


def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your account has been created! You are now able to log in')
            return redirect('login')
    else:
        form = UserRegisterForm()
    return render(request, 'tickets/register.html', {'form': form})


class CheckoutView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        cart = get_object_or_404(Cart, user=request.user)
        if not cart.cartitem_set.exists():
            messages.error(request, "Your cart is empty.")
            return redirect('cart_detail')

        total = sum(item.total_price() for item in cart.cartitem_set.all())
        payment_method = PaymentMethod.objects.last()
        form = PaymentProofForm()
        return render(request, 'tickets/checkout.html', {
            'cart': cart,
            'total': total,
            'form': form,
            'payment_method': payment_method
        })

    def post(self, request, *args, **kwargs):
        cart = get_object_or_404(Cart, user=request.user)
        total_str = request.POST.get('final_total')
        try:
            total = float(total_str) if total_str is not None else sum(item.total_price() for item in cart.cartitem_set.all())
        except ValueError:
            total = sum(item.total_price() for item in cart.cartitem_set.all())

        order = Order.objects.create(user=request.user)
        for cart_item in cart.cartitem_set.all():
            OrderItem.objects.create(
                order=order,
                ticket_category=cart_item.ticket_category,
                quantity=cart_item.quantity,
                price=cart_item.ticket_category.price,
            )
            ticket_category = cart_item.ticket_category
            ticket_category.booked_seats += cart_item.quantity
            ticket_category.reserved_seats -= cart_item.quantity
            ticket_category.save()

        cart.cartitem_set.all().delete()

        form = PaymentProofForm(request.POST, request.FILES)
        if form.is_valid():
            proof = form.save(commit=False)
            proof.order = order
            proof.price_paid = total
            proof.save()
            send_html_email(
                'Your Order Confirmation',
                'tickets/emails/payment_received.html',
                {
                    'user': request.user,
                    'cart': cart,
                    'order': order,
                    'total': int(total),
                },
                request.user.email
            )

            if request.POST.get('discount_code_form') == "True":
                code = request.POST.get('discount_code_name')
                discount_code = get_object_or_404(DiscountCode, code=code)
                if discount_code.is_valid():
                    discount_code.used_count += 1
                    discount_code.save()

        return redirect('order_confirmed')


def apply_discount(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        code = data.get('discount_code', '')
        cart = get_object_or_404(Cart, user=request.user)
        total = sum(item.total_price() for item in cart.cartitem_set.all())
        try:
            discount = DiscountCode.objects.get(code=code)
            if discount.is_valid():
                new_total = round(discount.apply_discount(total))
                formatted_total = "Rp " + f"{int(new_total):,}".replace(",", ".")
                return JsonResponse({
                    'success': True,
                    'new_total_formatted': formatted_total,
                    'new_total_numeric': str(new_total)
                })
            return JsonResponse({'success': False, 'message': 'Invalid or expired discount code.'})
        except DiscountCode.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invalid discount code.'})
    return JsonResponse({'success': False, 'message': 'Invalid request.'})


@staff_member_required
def admin_dashboard(request):
    orders = Order.objects.all().order_by('-created_at')
    seminars = Seminar.objects.all().order_by('date')

    for order in orders:
        order.aggregated_items = aggregate_order_items(order)
    return render(request, 'tickets/admin_dashboard.html', {'orders': orders, 'seminars': seminars})


@staff_member_required
def scicom_dashboard(request):
    scicom = SciComSubmission.objects.all()
    return render(request, 'tickets/scicom_dashboard.html', {'scicom': scicom})


@login_required
def profile_view(request):
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'profile.html', {'orders': orders})


@login_required
def order_history(request):
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    for order in orders:
        order.aggregated_items = aggregate_order_items(order)
    return render(request, 'tickets/order_history.html', {'orders': orders})


class SeminarDetailView(SeminarPaginationMixin, DetailView):
    model = Seminar
    template_name = 'tickets/ticket_detail.html'
    context_object_name = 'seminar'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ticket_categories'] = self.get_object().ticket_categories.all()
        search = self.request.GET.get('search', '')
        seminars = Seminar.objects.filter(title__icontains=search).order_by('date') if search \
            else Seminar.objects.all().order_by('date')
        context.update(self.paginate_seminars(seminars))
        return context


@login_required
def cart_item_count(request):
    cart, _ = Cart.objects.get_or_create(user=request.user)
    item_count = cart.cartitem_set.aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
    return JsonResponse({'item_count': item_count})


class CartDetailView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        cart, _ = Cart.objects.get_or_create(user=request.user)
        cart_items = cart.cartitem_set.all()
        return render(request, 'tickets/cart_detail.html', {
            'cart': cart,
            'is_empty': not cart_items.exists(),
            'cart_items': cart_items
        })


class RemoveFromCartView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        cart_item = get_object_or_404(CartItem, id=kwargs['item_id'])
        cart_item.delete()
        return redirect('cart_detail')


class AddToCartView(View):
    def post(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({'success': False, 'message': 'not_authenticated'})

        ticket_category = get_object_or_404(TicketCategory, id=request.POST.get('ticket_category_id'))
        cart, _ = Cart.objects.get_or_create(user=request.user)
        cart_item, created = CartItem.objects.get_or_create(
            cart=cart, ticket_category=ticket_category, defaults={'quantity': 0}
        )
        quantity = int(request.POST.get('quantity', 1))
        if ticket_category.remaining_seats < quantity:
            return JsonResponse({'success': False, 'message': 'Not enough remaining seats'})
        cart_item.quantity = (cart_item.quantity + quantity) if not created else quantity
        cart_item.save()
        return JsonResponse({
            'success': True,
            'message': 'Added to cart successfully!',
            'quantity': cart_item.quantity,
            'remaining_seats': ticket_category.remaining_seats,
            'item_count': cart.cartitem_set.count()
        })


@staff_member_required
@require_POST
def confirm_order_view(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    order.is_confirmed = True
    order.confirmation_date = timezone.now()
    order.transaction_id = request.POST.get('transaction_id')
    order.save()
    try:
        send_html_email(
            'Konfirmasi Pesanan Anda',
            'tickets/emails/order_confirmed.html',
            {'user': order.user},
            order.user.email
        )
    except Exception as e:
        messages.error(request, f"Failed to send email: {e}")
    return redirect('admin_dashboard')


@staff_member_required
def export_orders_view(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    headers = [
        'Username', 'Email', 'Nama Lengkap', 'NIK', 'NPWP', 'Institution',
        'No. Telfon', 'Order ID', 'Created At', 'Confirmed',
        'Confirmation Date', 'Seminars', 'Total Price'
    ]
    ws.append(headers)
    for order in Order.objects.all():
        created_at = order.created_at.replace(tzinfo=None) if order.created_at else None
        confirmation_date = order.confirmation_date.replace(tzinfo=None) if order.confirmation_date else None
        seminars_str = aggregate_order_items_str(order)
        proof = PaymentProof.objects.filter(order=order).first()
        price_paid = float(proof.price_paid) if proof else 0.0
        user = order.user
        ws.append([
            user.username,
            user.email,
            user.nama_lengkap,
            user.nik,
            user.npwp,
            user.institution,
            user.Nomor_telpon,
            order.id,
            created_at,
            'Yes' if order.is_confirmed else 'No',
            confirmation_date,
            seminars_str,
            price_paid,
        ])
    for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
        for cell in row:
            cell.number_format = '#,##0'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="order_report.xlsx"'
    wb.save(response)
    return response


@staff_member_required
def export_orders_for_seminar_view(request, seminar_id):
    seminar = get_object_or_404(Seminar, pk=seminar_id)
    orders = Order.objects.filter(orderitem__ticket_category__seminar=seminar).distinct()
    wb = Workbook()
    ws = wb.active
    sanitized_title = re.sub(r'[\/:*?"<>|]', '', f"Orders for {seminar.title}")
    ws.title = sanitized_title
    headers = [
        'Username', 'Email', 'Nama Lengkap', 'NIK', 'NPWP', 'Institution',
        'No. Telfon', 'Order ID', 'Created At', 'Confirmed',
        'Confirmation Date', 'Seminars', 'Total Price'
    ]
    ws.append(headers)
    for order in orders:
        created_at = order.created_at.replace(tzinfo=None) if order.created_at else None
        confirmation_date = order.confirmation_date.replace(tzinfo=None) if order.confirmation_date else None
        seminars_str = aggregate_order_items_str(order)
        proof = PaymentProof.objects.filter(order=order).first()
        price_paid = float(proof.price_paid) if proof else 0.0
        user = order.user
        ws.append([
            user.username,
            user.email,
            user.nama_lengkap,
            user.nik,
            user.npwp,
            user.institution,
            user.Nomor_telpon,
            order.id,
            created_at,
            'Yes' if order.is_confirmed else 'No',
            confirmation_date,
            seminars_str,
            price_paid,
        ])
    for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
        for cell in row:
            cell.number_format = '#,##0'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"orders_for_{seminar.title.replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_scicom_submissions_excel(request):
    workbook = Workbook()
    abstracts = SciComSubmission.objects.filter(submission_type=SciComSubmission.ABSTRACT).select_related('user')
    videos = SciComSubmission.objects.filter(submission_type=SciComSubmission.VIDEO).select_related('user')
    flyers = SciComSubmission.objects.filter(submission_type=SciComSubmission.FLYER).select_related('user')

    # ABSTRACT Submissions
    abstract_ws = workbook.active
    abstract_ws.title = "Abstract Submissions"
    abstract_columns = [
        "ID", "Name", "Occupation", "Email", "Phone", "Affiliation", "Address",
        "Already Registered?", "Created At", "Abstract Title", "Paper Type", "Abstract Authors",
        "Abstract Text", "Link Abstract",
    ]
    abstract_ws.append(abstract_columns)
    for submission in abstracts:
        abstract_ws.append([
            submission.id,
            submission.name,
            submission.get_occupation_display(),
            submission.email,
            submission.phone,
            submission.affiliation,
            submission.address,
            "Yes" if submission.already_registered else "No",
            submission.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            submission.abstract_title,
            submission.paper_type,
            submission.abstract_authors,
            submission.abstract_text,
            submission.link_abstract,
        ])

    # VIDEO Submissions
    video_ws = workbook.create_sheet("Video Submissions")
    video_columns = [
        "ID", "Name", "Occupation", "Email", "Phone", "Affiliation", "Address",
        "Already Registered?", "Created At", "Video Title", "Video Authors", "Link Video",
    ]
    video_ws.append(video_columns)
    for submission in videos:
        video_ws.append([
            submission.id,
            submission.name,
            submission.get_occupation_display(),
            submission.email,
            submission.phone,
            submission.affiliation,
            submission.address,
            "Yes" if submission.already_registered else "No",
            submission.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            submission.video_title,
            submission.video_authors,
            submission.link_video,
        ])

    # FLYER Submissions
    flyer_ws = workbook.create_sheet("Flyer Submissions")
    flyer_columns = [
        "ID", "Name", "Occupation", "Email", "Phone", "Affiliation", "Address",
        "Already Registered?", "Created At", "Flyer Title", "Flyer Authors", "Link Flyer",
    ]
    flyer_ws.append(flyer_columns)
    for submission in flyers:
        flyer_ws.append([
            submission.id,
            submission.name,
            submission.get_occupation_display(),
            submission.email,
            submission.phone,
            submission.affiliation,
            submission.address,
            "Yes" if submission.already_registered else "No",
            submission.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            submission.flyer_title,
            submission.flyer_authors,
            submission.link_flyer,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = "scicom_submissions_multi_" + datetime.datetime.now().strftime("%Y%m%d") + ".xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response
